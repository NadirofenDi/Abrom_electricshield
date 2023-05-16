from PIL import Image
from pathlib import Path
import openpyxl

class Result:
    pass


x_allowed = 250
y_allowed = 150

workbook = openpyxl.load_workbook('XLS file/Electric_shield.xlsx')
list_images = []

image = Image.new('RGB', (2800, 5200), color='white')

borders = Image.open('app/models/borders.png')
image.paste(borders, (0, 0), mask=borders)

klemma2 = Image.open('app/models/0.png')
klemma2 = klemma2.crop((1001, 829, 1240, 1440))
klemma2_path = Path("app/models_named/klemma2.png")
list_images.append(klemma2_path.name.replace('.png', ''))
width_klemma2, height_klemma2 = klemma2.size

#
# sheet = workbook['Sheet1']
# for row in sheet.iter_rows(values_only=True):
#     if row[0] in list_images:
#         sum = 0
#         for i in range(row[1]):
#             sum += width_klemma2
#             if sum >= 2500:
#                 print('Не достаточно места в ряду')
#                 break
#             else:


########################################

WB_MR6C_v2 = Image.open('app/models/1.png')
WB_MR6C_v2 = WB_MR6C_v2.crop((382, 744, 793, 1632))
WB_MR6C_v2_path = Path("app/models_named/WB_MR6C_v2.png")
list_images.append(WB_MR6C_v2_path.name.replace('.png', ''))
width_WB_MR6C_v2, height_WB_MR6C_v2 = WB_MR6C_v2.size


# sheet = workbook['Sheet1']
# for row in sheet.iter_rows(values_only=True):
#     print(row)
#     if row[0] in list_images:
#         sum = 0
#         for i in range(row[1]):
#             sum += width_WB_MR6C_v2
#             if sum >= 2500:
#                 print('Не достаточно места в ряду')
#                 break
#             else:


image.paste(WB_MR6C_v2, (0, 0), mask=WB_MR6C_v2)
x_allowed += width_WB_MR6C_v2
image.paste(klemma2, (x_allowed, 150), mask=klemma2)



print(list_images)
image.show()
